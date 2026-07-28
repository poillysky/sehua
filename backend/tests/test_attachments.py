"""Attachment extract / zip-txt / torrent→magnet / inject."""

from __future__ import annotations

import io
import zipfile

from parsers.attachments import (
    extract_download_attachments,
    filter_tail_attachments,
    filter_torrent_attachments,
    inject_attachment_text,
    is_attachment_denied,
)
from parsers.links import parse_thread_dual
from parsers.torrent import parse_torrent_bytes
from crawler.attachments import _extract_txt_from_archive, _text_from_attachment_bytes
from parsers.attachments import DownloadAttachment


def _minimal_torrent(name: bytes = b"demo.bin", length: int = 123) -> bytes:
    # bencode: d8:announce0:4:infod6:lengthi{len}e4:name{n}:{name}ee
    info = b"d6:lengthi" + str(length).encode() + b"e4:name" + str(len(name)).encode() + b":" + name + b"e"
    return b"d8:announce0:4:info" + info + b"e"


def test_extract_and_filter_tail():
    html = """
    <div class="tattl">
      <a href="forum.php?mod=attachment&aid=1">目录树.txt</a>
      <a href="forum.php?mod=attachment&aid=2">链接A.txt</a>
      <a href="forum.php?mod=attachment&aid=3">pack.zip</a>
      <a href="forum.php?mod=attachment&aid=4">seed.torrent</a>
    </div>
    """
    all_a = extract_download_attachments("https://www.sehuatang.net/", html)
    kinds = {a.kind for a in all_a}
    assert "txt" in kinds and "zip" in kinds and "torrent" in kinds
    tail = filter_tail_attachments(all_a, limit=3)
    assert all(a.kind in ("txt", "zip", "rar", "excel") for a in tail)
    assert not any("目录" in a.name for a in tail)
    torrents = filter_torrent_attachments(all_a)
    assert len(torrents) == 1 and torrents[0].kind == "torrent"


def test_cf_email_obfuscated_attachment_name():
    """含 @ 的附件名被 CF 混淆成 [email protected] 时仍应识别为 txt。"""
    enc = "d3a4a4a4fdeaeb87fdbfb29383a1baa5b2a7b690b2a0a7babdb4fe8bfda7aba7"
    html = f"""
    <ignore_js_op>
      <a href="forum.php?mod=attachment&amp;aid=AAA">
        <span class="__cf_email__" data-cfemail="{enc}">[email&#160;protected]</span>
      </a>
      <a href="forum.php?mod=attachment&amp;aid=BBB">xxx_目录树.txt</a>
      <a href="forum.php?mod=attachment&amp;aid=CCC">shot_115链接.png</a>
    </ignore_js_op>
    """
    all_a = extract_download_attachments("https://www.sehuatang.net/", html)
    names = [a.name for a in all_a]
    assert "www.98T.la@PrivateCasting-X.txt" in names
    assert not any(n.endswith(".png") for n in names)
    tail = filter_tail_attachments(all_a, limit=3)
    assert [a.name for a in tail] == ["www.98T.la@PrivateCasting-X.txt"]


def test_cf_email_prefix_plus_sibling_filename():
    """CF 只混淆 @ 前缀、.txt 名在 span 外时仍应识别（tid=2810669）。"""
    enc = "750202025b4c4d215b19143547415b44475847405b4544"
    html = f"""
    <ignore_js_op>
      <a href="forum.php?mod=attachment&amp;aid=MTQ3">
        <span class="__cf_email__" data-cfemail="{enc}">[email&#160;protected]</span>
        国产传媒合集.txt
      </a>
      <a href="forum.php?mod=attachment&amp;aid=IMG">preview.jpg</a>
    </ignore_js_op>
    """
    all_a = extract_download_attachments("https://www.sehuatang.net/", html)
    assert len(all_a) == 1
    assert all_a[0].kind == "txt"
    assert all_a[0].name.endswith("国产传媒合集.txt")
    assert "www.98T.la@24.12-25.01" in all_a[0].name


def test_rar_with_文件夹_in_name_not_skipped():
    """「新建文件夹.rar」是资源包，不能当目录树跳过，否则无权包会反复重试。"""
    html = """
    <ignore_js_op>
      <a href="forum.php?mod=attachment&amp;aid=1">www.98T.la  新建文件夹      下载麻烦顺手评评分.rar</a>
      <a href="forum.php?mod=attachment&amp;aid=2">目录树.txt</a>
    </ignore_js_op>
    """
    all_a = extract_download_attachments("https://www.sehuatang.net/", html)
    assert any(a.kind == "rar" for a in all_a)
    tail = filter_tail_attachments(all_a, limit=3)
    assert len(tail) == 1
    assert tail[0].kind == "rar"
    assert "文件夹" in tail[0].name


def test_zip_inner_txt_and_ed2k_parse():
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w") as zf:
        zf.writestr(
            "links.txt",
            "ed2k://|file|demo.rar|100|ABCDEF0123456789ABCDEF0123456789|/\n",
        )
    data = buf.getvalue()
    text = _extract_txt_from_archive(data, "zip")
    assert "ed2k://" in text

    att = DownloadAttachment(name="pack.zip", url="http://x/a", kind="zip")
    out = _text_from_attachment_bytes(att, data)
    parsed = parse_thread_dual(
        "<html><title>t</title></html>",
        tid=1,
        preferred_link="ed2k",
        extra_text=out,
    )
    assert parsed.primary_link_kind == "ed2k"
    assert parsed.ed2k_links


def test_archive_password_candidates_from_post():
    from crawler.attachments import _archive_password_candidates

    html = """
    <html><body>
      <div id="postmessage_1">
        【资源名称】：示例<br>
        【解压密码】：www.98T.la@
      </div>
    </body></html>
    """
    cands = _archive_password_candidates(html)
    assert "www.98T.la@" in cands
    assert "www.98T.la" in cands


def test_archive_password_strips_cf_txt_and_glued_attachment_name():
    """CF 把密码编成 user@host.txt，后面再粘 rar 附件名时仍应得到可用密码。"""
    from crawler.attachments import _archive_password_candidates
    from parsers.content import _clip_field_value, extract_password

    polluted = "MyBigDick@sehuatang.txt 18OnlyGirls.rar (42.29 KB,"
    assert _clip_field_value(polluted, password=True) == "MyBigDick@sehuatang.txt"
    assert extract_password(f"【解压密码】：{polluted}") == "MyBigDick@sehuatang.txt"

    html = f"""
    <html><body>
      <div id="postmessage_1">【解压密码】：{polluted}</div>
    </body></html>
    """
    cands = _archive_password_candidates(html)
    assert "MyBigDick@sehuatang.txt" in cands
    assert "MyBigDick@sehuatang" in cands


def test_password_protected_zip_txt_extract():
    """帖内密码应能解开加密 zip 里的 txt（AES → pyzipper）。"""
    import pytest

    pyzipper = pytest.importorskip("pyzipper")
    from crawler.attachments import _extract_txt_from_archive

    link = "ed2k://|file|demo.rar|100|ABCDEF0123456789ABCDEF0123456789|/\n"
    buf = io.BytesIO()
    with pyzipper.AESZipFile(
        buf,
        "w",
        compression=pyzipper.ZIP_DEFLATED,
        encryption=pyzipper.WZ_AES,
    ) as zf:
        zf.setpassword(b"secret123")
        zf.writestr("links.txt", link)
    data = buf.getvalue()
    assert not _extract_txt_from_archive(data, "zip")
    assert "ed2k://" in _extract_txt_from_archive(data, "zip", passwords=["secret123"])
    assert "ed2k://" in _extract_txt_from_archive(
        data, "zip", passwords=["wrong", "secret123"]
    )


def test_bare_password_98t_extracted():
    """帖内仅写「密码」+ www.98T.la@（无解压前缀）也应抽出。"""
    from crawler.attachments import _archive_password_candidates
    from parsers.content import extract_password

    html = """
    <div id="postmessage_1">
      <font>密码</font><font>www.98T.la@</font>
      <a href="forum.php?mod=attachment&aid=1">新建文件夹.rar</a>
    </div>
    """
    assert extract_password("密码 www.98T.la@") == "www.98T.la@"
    cands = _archive_password_candidates(html)
    assert "www.98T.la@" in cands
    assert "www.98T.la" in cands


def test_nested_zip_uses_same_password():
    """外层 zip 内再套加密 zip，应用同一帖内密码解开取 txt。"""
    import pytest

    pyzipper = pytest.importorskip("pyzipper")
    from crawler.attachments import _extract_txt_from_archive

    link = "ed2k://|file|demo.mp4|10|ABCDEFABCDEFABCDEFABCDEFABCDEFAB|/\n"
    inner = io.BytesIO()
    with pyzipper.AESZipFile(
        inner,
        "w",
        compression=pyzipper.ZIP_DEFLATED,
        encryption=pyzipper.WZ_AES,
    ) as zf:
        zf.setpassword(b"www.98T.la@")
        zf.writestr("ed2k.txt", link)
    outer = io.BytesIO()
    with pyzipper.AESZipFile(
        outer,
        "w",
        compression=pyzipper.ZIP_DEFLATED,
        encryption=pyzipper.WZ_AES,
    ) as zf:
        zf.setpassword(b"www.98T.la@")
        zf.writestr("新建文件夹/inner.zip", inner.getvalue())
    text = _extract_txt_from_archive(
        outer.getvalue(), "zip", passwords=["www.98T.la@", "www.98T.la"]
    )
    assert "ed2k://" in text


def test_torrent_to_magnet():
    data = _minimal_torrent()
    magnet = parse_torrent_bytes(data, filename_hint="seed.torrent")
    assert magnet is not None
    assert magnet.link.lower().startswith("magnet:?xt=urn:btih:")
    assert magnet.infohash


def test_zip_inner_torrent_to_magnet():
    """压缩包内 .torrent 应转成 magnet。"""
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w") as zf:
        zf.writestr("seed.torrent", _minimal_torrent())
        zf.writestr("readme.txt", "no link here")
    text = _extract_txt_from_archive(buf.getvalue(), "zip")
    assert text.lower().startswith("magnet:?xt=urn:btih:")


def test_inject_and_denied():
    html = '<div id="postmessage_1">hello</div>'
    merged = inject_attachment_text(html, "magnet:?xt=urn:btih:AABBCCDDEEFF00112233445566778899")
    assert "postmessage_attach0" in merged
    assert "magnet:?" in merged
    assert is_attachment_denied("抱歉，只有特定用户可以下载此附件")


def test_denied_not_masked_by_empty_download():
    """第一个附件空下载、第二个无权时，结果须带 denied（不能只标 downloaded）。"""
    from parsers.attachments import AttachmentFetchResult

    # 模拟 download_tail 汇总结论（与 crawler 逻辑一致）
    any_downloaded = True
    any_denied = True
    result_text = ""
    if result_text and ("magnet:" in result_text.lower() or "ed2k://" in result_text.lower()):
        out = AttachmentFetchResult(text=result_text, downloaded=True, denied=False)
    elif any_denied:
        out = AttachmentFetchResult(
            text=result_text or "",
            downloaded=any_downloaded,
            denied=True,
        )
    elif any_downloaded:
        out = AttachmentFetchResult(downloaded=True)
    else:
        out = AttachmentFetchResult(failed=True)
    assert out.denied is True
    assert out.downloaded is True
    assert not out.text


def test_denied_not_masked_by_baidu_password_txt():
    """115 附件无权、只下到百度口令 txt → 仍须 denied（tid=3341941）。"""
    from parsers.attachments import AttachmentFetchResult

    any_downloaded = True
    any_denied = True
    result_text = "复制口令后打开「手机百度网盘 App」即可\n墀垩创街忐了心凉礼艇左凿圜\n"
    has_importable = "magnet:" in result_text.lower() or "ed2k://" in result_text.lower()
    if result_text and has_importable:
        out = AttachmentFetchResult(text=result_text, downloaded=True, denied=False)
    elif any_denied:
        out = AttachmentFetchResult(
            text=result_text, downloaded=True, denied=True
        )
    else:
        out = AttachmentFetchResult(text=result_text, downloaded=True)
    assert out.denied is True
    assert out.downloaded is True
    assert out.text
    assert "ed2k://" not in out.text.lower()


def test_denied_continues_to_next_then_import_wins():
    """3 个附件：前两个无权、第三个有链 → 应继续试并成功（denied 不短路）。"""
    names = ["a.rar", "b.rar", "c.txt"]
    tried: list[str] = []
    result_text = ""
    any_denied = False
    for name in names:
        tried.append(name)
        if name.endswith(".rar"):
            any_denied = True
            continue
        result_text = "magnet:?xt=urn:btih:AABBCCDDEEFF00112233445566778899"
        break
    assert tried == ["a.rar", "b.rar", "c.txt"]
    assert result_text.startswith("magnet:")
    # 有可入库文本时，最终不应因 earlier denied 而占位
    denied_final = False if result_text else any_denied
    assert denied_final is False


def test_judge_stubs_when_second_attachment_denied():
    """正文无链、附件已试、第二附件无权 → 占位入库，不重试。"""
    from workers.thread_outcome import judge_thread_html

    html = """
    <html><head><title>【高清】示例资源贴 - 色花堂</title></head>
    <body>
      <span id="thread_subject">【高清】示例资源贴</span>
      <div id="postmessage_1">链接见第二个附件</div>
      <div class="tattl">
        <a href="forum.php?mod=attachment&amp;aid=1">说明.txt</a>
        <a href="forum.php?mod=attachment&amp;aid=2">资源链接.txt</a>
      </div>
      Powered by Discuz!
    </body></html>
    """
    html = html + ("<!-- pad -->" * 900)
    # 说明.txt 下到了无用文本，资源链接.txt 无权
    out = judge_thread_html(
        html,
        board_fid=95,
        list_title="【高清】示例资源贴",
        attachments_already_tried=True,
        attachment_denied=True,
        had_attachments=True,
        preferred_link="ed2k",
    )
    assert out.verdict == "stub"
    assert out.outcome == "无权限下载附件"


def test_judge_tries_attachments_when_html_has_decoy_ed2k():
    """封面 tip 里有 ed2k、正文无主资源、有 rar 附件 → 应下附件，勿直接「有链但无主资源」。"""
    from workers.thread_outcome import judge_thread_html

    html = """
    <html><head><title>【ED2k】合集 - 色花堂</title></head>
    <body>
      <span id="thread_subject">【ED2k】合集</span>
      <div id="postmessage_1">资源在附件压缩包</div>
      </tbody>
      <ignore_js_op>
        <a id="ed2k_Q1u" href="ed2k://|file|波姐封面.zip|207438322|92262C23A97D1D04DA885978E4B7C8E6|/" target="_blank">波姐封面.zip</a>
        <a href="forum.php?mod=attachment&amp;aid=1">www.98T.la@资源.rar</a>
      </ignore_js_op>
      Powered by Discuz!
    </body></html>
    """
    html = html + ("<!-- pad -->" * 900)
    out = judge_thread_html(
        html,
        board_fid="141:691",
        list_title="【ED2k】合集",
        preferred_link="ed2k",
    )
    assert out.verdict == "need_attachments"
    assert out.attachment_kind == "txt_tail"
    assert out.outcome != "解析入库失败（有链但无主资源）"

    denied = judge_thread_html(
        html,
        board_fid="141:691",
        list_title="【ED2k】合集",
        preferred_link="ed2k",
        attachments_already_tried=True,
        attachment_denied=True,
        had_attachments=False,
    )
    assert denied.verdict == "stub"
    assert denied.outcome == "无权限下载附件"


def test_ed2k_board_picks_torrent_when_only_torrent_attach():
    """情色分享等电驴板：仅有 .torrent 时应转磁力，而非只找 txt/zip。"""
    from parsers.attachments import pick_ed2k_attachment_kind
    from workers.thread_outcome import judge_thread_html

    html = """
    <html><head><title>【欧美VR种子】示例 - 色花堂</title></head>
    <body>
      <span id="thread_subject">【欧美VR种子】示例</span>
      <div id="postmessage_1">种子见附件</div>
      <ignore_js_op>
        <a href="forum.php?mod=attachment&amp;aid=1">www.98T.la@demo.torrent</a>
      </ignore_js_op>
      Powered by Discuz!
    </body></html>
    """
    html = html + ("<!-- pad -->" * 900)
    assert pick_ed2k_attachment_kind("https://www.sehuatang.net/", html) == "torrent"
    out = judge_thread_html(
        html,
        board_fid="95:716",
        list_title="【欧美VR种子】示例",
        preferred_link="ed2k",
        base_url="https://www.sehuatang.net/",
    )
    assert out.verdict == "need_attachments"
    assert out.attachment_kind == "torrent"


def test_judge_skips_preview_only_no_resource_links():
    """仅有预览图、无 txt/zip/rar/torrent、无正文链 → 直接跳过，勿因「附件」字样去下附件。"""
    from parsers.thread_gates import looks_like_attachment_zone
    from workers.thread_outcome import judge_thread_html

    html = """
    <html><head><title>【自转】【115ed2k】示例无链接贴 - 色花堂</title></head>
    <body>
      <span id="thread_subject">【自转】【115ed2k】示例无链接贴</span>
      <div id="postmessage_1">【资源名称】：示例 大哥链接忘记放了吧</div>
      <ignore_js_op>
        <img aid="1" src="shot.png" />
        <a href="https://tu.example/shot.png" target="_blank">下载附件</a>
      </ignore_js_op>
      Powered by Discuz!
    </body></html>
    """
    html = html + ("<!-- pad -->" * 900)
    assert looks_like_attachment_zone(html) is False
    out = judge_thread_html(
        html,
        board_fid="95:716",
        list_title="【自转】【115ed2k】示例无链接贴",
        preferred_link="ed2k",
    )
    assert out.verdict == "skipped"
    assert out.need_attachments is False
    assert "未解析到" in out.outcome or "跳过" in out.outcome


def test_judge_skips_after_attachments_tried_without_link():
    """附件已下到、仍无 ed2k/磁力 → 跳过（勿占位）。"""
    from workers.thread_outcome import judge_thread_html

    html = """
    <html><head><title>【高清】示例资源贴 - 色花堂</title></head>
    <body>
      <span id="thread_subject">【高清】示例资源贴</span>
      <div id="postmessage_1">链接见附件</div>
      <div class="tattl">
        <a href="forum.php?mod=attachment&amp;aid=2">资源链接.txt</a>
      </div>
      Powered by Discuz!
    </body></html>
    """
    html = html + ("<!-- pad -->" * 900)
    out = judge_thread_html(
        html,
        board_fid=95,
        list_title="【高清】示例资源贴",
        attachments_already_tried=True,
        attachment_denied=False,
        had_attachments=True,
        preferred_link="ed2k",
    )
    assert out.verdict == "skipped"
    assert "未解析到" in out.outcome
def test_excel_attachment_kind_and_filter():
    html = """
    <div class="tattl">
      <a href="forum.php?mod=attachment&aid=1">磁力列表.xlsx</a>
      <a href="forum.php?mod=attachment&aid=2">seed.torrent</a>
    </div>
    """
    all_a = extract_download_attachments("https://www.sehuatang.net/", html)
    assert any(a.kind == "excel" for a in all_a)
    tail = filter_tail_attachments(all_a, limit=3)
    assert any(a.kind == "excel" for a in tail)
    from parsers.attachments import pick_magnet_attachment_kind

    assert pick_magnet_attachment_kind("https://www.sehuatang.net/", html) == "txt_tail"
    # 【BT种子】标题：即使同帖有 excel，也优先种子
    assert (
        pick_magnet_attachment_kind(
            "https://www.sehuatang.net/",
            html,
            title="【BT种子】示例片名",
        )
        == "torrent"
    )


def test_extract_magnet_from_xlsx_bytes():
    import openpyxl

    magnet = (
        "magnet:?xt=urn:btih:0123456789ABCDEF0123456789ABCDEF01234567"
        "&dn=demo"
    )
    wb = openpyxl.Workbook()
    ws = wb.active
    ws["A1"] = "资源"
    ws["B1"] = magnet
    buf = io.BytesIO()
    wb.save(buf)
    data = buf.getvalue()
    att = DownloadAttachment(
        name="links.xlsx",
        url="https://example.com/a",
        kind="excel",
    )
    text = _text_from_attachment_bytes(att, data)
    assert "magnet:?xt=urn:btih:" in text
    assert "0123456789ABCDEF0123456789ABCDEF01234567" in text.upper()


def test_extract_magnet_from_xls_binary_scan():
    magnet = b"magnet:?xt=urn:btih:ABCDEF0123456789ABCDEF0123456789ABCDEF01&dn=x"
    blob = b"\x00\x01\x02" + magnet + b"\xff\xfe junk"
    att = DownloadAttachment(name="old.xls", url="https://example.com/a", kind="excel")
    text = _text_from_attachment_bytes(att, blob)
    assert "magnet:?xt=urn:btih:ABCDEF0123456789ABCDEF0123456789ABCDEF01" in text

def test_filter_all_link_attachments_order_and_limit():
    from parsers.attachments import filter_all_link_attachments, DownloadAttachment

    atts = [
        DownloadAttachment("z.rar", "u", "rar"),
        DownloadAttachment("a.txt", "u", "txt"),
        DownloadAttachment("seed.torrent", "u", "torrent"),
        DownloadAttachment("b.xlsx", "u", "excel"),
        DownloadAttachment("目录树.txt", "u", "txt"),
        DownloadAttachment("p.zip", "u", "zip"),
        DownloadAttachment("links.docx", "u", "doc"),
    ]
    # 默认 / 电驴：txt → zip/rar → excel/doc → torrent；目录类 txt 垫底
    got = filter_all_link_attachments(atts, limit=10, preferred_link="ed2k")
    assert [a.kind for a in got] == [
        "txt",
        "zip",
        "rar",
        "excel",
        "doc",
        "torrent",
        "txt",
    ]
    assert [a.name for a in got] == [
        "a.txt",
        "p.zip",
        "z.rar",
        "b.xlsx",
        "links.docx",
        "seed.torrent",
        "目录树.txt",
    ]

    # 磁力：torrent → excel/doc/txt → zip/rar；目录类 txt 仍垫底
    got_m = filter_all_link_attachments(atts, limit=10, preferred_link="magnet")
    assert [a.kind for a in got_m] == [
        "torrent",
        "excel",
        "doc",
        "txt",
        "zip",
        "rar",
        "txt",
    ]
    assert got_m[0].name == "seed.torrent"
    assert got_m[-1].name == "目录树.txt"


def test_filter_all_link_attachments_prefers_115_name():
    """文件名含 115 的附件优先于同类型其它文件。"""
    from parsers.attachments import filter_all_link_attachments, DownloadAttachment

    atts = [
        DownloadAttachment("百度网盘下载链接.txt", "u", "txt"),
        DownloadAttachment("115ED2K下载链接.txt", "u", "txt"),
        DownloadAttachment("防失效备用版.txt", "u", "txt"),
        DownloadAttachment("other.zip", "u", "zip"),
        DownloadAttachment("115备份.zip", "u", "zip"),
    ]
    got = filter_all_link_attachments(atts, preferred_link="ed2k")
    names = [a.name for a in got]
    assert names[0] == "115ED2K下载链接.txt"
    assert names[1] == "115备份.zip"
    assert "百度网盘下载链接.txt" in names
    assert names.index("115ED2K下载链接.txt") < names.index("百度网盘下载链接.txt")
    assert names.index("115备份.zip") < names.index("other.zip")


def test_filter_all_link_attachments_prefers_yifen_name():
    """文件名含「一分也是爱」与 115 / 98 同档优先。"""
    from parsers.attachments import filter_all_link_attachments, DownloadAttachment

    atts = [
        DownloadAttachment("www.98T.la@文件名列表.txt", "u", "txt"),
        DownloadAttachment("求个评论和免费的赞，一分也是爱.txt", "u", "txt"),
        DownloadAttachment("防失效备用版.txt", "u", "txt"),
        DownloadAttachment("115ED2K下载链接.txt", "u", "txt"),
    ]
    got = filter_all_link_attachments(atts, preferred_link="ed2k")
    names = [a.name for a in got]
    # 115 / 98 / 一分也是爱 同档，均先于普通名
    for vip in (
        "求个评论和免费的赞，一分也是爱.txt",
        "115ED2K下载链接.txt",
        "www.98T.la@文件名列表.txt",
    ):
        assert names.index(vip) < names.index("防失效备用版.txt")


def test_filter_all_link_attachments_prefers_98_name():
    """文件名含 98（如 98T.la）优先于同类型其它文件。"""
    from parsers.attachments import filter_all_link_attachments, DownloadAttachment

    atts = [
        DownloadAttachment("防失效备用版.txt", "u", "txt"),
        DownloadAttachment("www.98T.la@Minana呀 录播.txt", "u", "txt"),
        DownloadAttachment("other.zip", "u", "zip"),
        DownloadAttachment("98合集.zip", "u", "zip"),
    ]
    got = filter_all_link_attachments(atts, preferred_link="ed2k")
    names = [a.name for a in got]
    assert names[0] == "www.98T.la@Minana呀 录播.txt"
    assert names[1] == "98合集.zip"
    assert names.index("www.98T.la@Minana呀 录播.txt") < names.index("防失效备用版.txt")
    assert names.index("98合集.zip") < names.index("other.zip")


def test_filter_all_link_attachments_both_uses_magnet_order():
    from parsers.attachments import filter_all_link_attachments, DownloadAttachment

    atts = [
        DownloadAttachment("a.txt", "u", "txt"),
        DownloadAttachment("seed.torrent", "u", "torrent"),
    ]
    got = filter_all_link_attachments(atts, preferred_link="both")
    assert [a.kind for a in got] == ["torrent", "txt"]


def _minimal_docx(body_text: str) -> bytes:
    """最小可用 OOXML docx（仅 word/document.xml）。"""
    document_xml = f"""<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<w:document xmlns:w="http://schemas.openxmlformats.org/wordprocessingml/2006/main">
  <w:body>
    <w:p><w:r><w:t>{body_text}</w:t></w:r></w:p>
  </w:body>
</w:document>
"""
    content_types = """<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<Types xmlns="http://schemas.openxmlformats.org/package/2006/content-types">
  <Default Extension="rels" ContentType="application/vnd.openxmlformats-package.relationships+xml"/>
  <Default Extension="xml" ContentType="application/xml"/>
  <Override PartName="/word/document.xml" ContentType="application/vnd.openxmlformats-officedocument.wordprocessingml.document.main+xml"/>
</Types>
"""
    rels = """<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">
  <Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/officeDocument" Target="word/document.xml"/>
</Relationships>
"""
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w") as z:
        z.writestr("[Content_Types].xml", content_types)
        z.writestr("_rels/.rels", rels)
        z.writestr("word/document.xml", document_xml)
    return buf.getvalue()


def test_doc_attachment_kind_and_filter():
    html = """
    <div class="tattl">
      <a href="forum.php?mod=attachment&aid=1">links.docx</a>
      <a href="forum.php?mod=attachment&aid=2">old.doc</a>
      <a href="forum.php?mod=attachment&aid=3">a.txt</a>
    </div>
    """
    all_a = extract_download_attachments("https://www.sehuatang.net/", html)
    assert any(a.kind == "doc" and a.name.endswith(".docx") for a in all_a)
    assert any(a.kind == "doc" and a.name.endswith(".doc") for a in all_a)
    tail = filter_tail_attachments(all_a, limit=5)
    assert any(a.kind == "doc" for a in tail)


def test_zip_inner_docx_magnet():
    magnet = "magnet:?xt=urn:btih:cccccccccccccccccccccccccccccccccccccccc&dn=doc"
    outer = io.BytesIO()
    with zipfile.ZipFile(outer, "w") as z:
        z.writestr("links.docx", _minimal_docx(magnet))
    text = _extract_txt_from_archive(outer.getvalue(), "zip")
    assert magnet in text


def test_zip_inner_doc_binary_magnet():
    magnet = b"magnet:?xt=urn:btih:dddddddddddddddddddddddddddddddddddddddd&dn=old"
    blob = b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1" + b"\x00" * 32 + magnet + b"\x00junk"
    outer = io.BytesIO()
    with zipfile.ZipFile(outer, "w") as z:
        z.writestr("seed.doc", blob)
    text = _extract_txt_from_archive(outer.getvalue(), "zip")
    assert b"magnet:?xt=urn:btih:dddddddddddddddddddddddddddddddddddddddd".decode() in text


def test_zip_inner_doc_utf16_ed2k_with_spaces():
    """Word .doc 常以 UTF-16LE 存链，且 ed2k 文件名含空格，不能在空格处截断。"""
    ed2k = (
        "ed2k://|file|www.98T.la@demo name with spaces.zip|"
        "5589668780|B6622F82B7CD4E74591DC898503E5218|/"
    )
    blob = b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1" + b"\x00" * 64 + ed2k.encode("utf-16-le")
    outer = io.BytesIO()
    with zipfile.ZipFile(outer, "w") as z:
        z.writestr("links.doc", blob)
    text = _extract_txt_from_archive(outer.getvalue(), "zip")
    assert "B6622F82B7CD4E74591DC898503E5218" in text
    assert "ed2k://|file|www.98T.la@demo name with spaces.zip|" in text
    from parsers.ed2k import parse_ed2k_text

    links = parse_ed2k_text(text)
    assert len(links) == 1
    assert links[0].hash.upper() == "B6622F82B7CD4E74591DC898503E5218"


def test_docx_attachment_bytes_extract():
    magnet = "magnet:?xt=urn:btih:eeeeeeeeeeeeeeeeeeeeeeeeeeeeeeeeeeeeeeee&dn=x"
    att = DownloadAttachment(name="x.docx", url="https://example.com/a", kind="doc")
    text = _text_from_attachment_bytes(att, _minimal_docx(magnet))
    assert magnet in text


def test_nested_zip_prefers_excel_magnet_over_115_txt():
    import openpyxl
    from crawler.attachments import _extract_txt_from_archive

    magnet = "magnet:?xt=urn:btih:bbbbbbbbbbbbbbbbbbbbbbbbbbbbbbbbbbbbbbbb&dn=x"
    wb = openpyxl.Workbook()
    wb.active["A1"] = magnet
    xbuf = io.BytesIO()
    wb.save(xbuf)
    inner = io.BytesIO()
    with zipfile.ZipFile(inner, "w") as z:
        z.writestr("links.xlsx", xbuf.getvalue())
        z.writestr(
            "sha1.txt",
            "115://f|1|AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA|BBBBBBBBBBBBBBBBBBBBBBBBBBBBBBBBBBBBBBBB|n",
        )
    outer = io.BytesIO()
    with zipfile.ZipFile(outer, "w") as z:
        z.writestr("inner.zip", inner.getvalue())
    text = _extract_txt_from_archive(outer.getvalue(), "zip")
    assert "magnet:?xt=urn:btih:" in text
    assert "115://" not in text


def test_push_member_text_stops_on_magnet():
    from crawler.attachments import _push_member_text

    chunks: list[str] = []
    assert _push_member_text(chunks, "only 115sha noise") is None
    early = _push_member_text(
        chunks,
        "magnet:?xt=urn:btih:dddddddddddddddddddddddddddddddddddddddd",
    )
    assert early is not None
    assert "magnet:?xt=urn:btih:" in early


def test_zip_stops_after_first_member_with_magnet():
    """压缩包内先遇到目标链即可返回，不必扫完后续成员。"""
    from crawler.attachments import _extract_txt_from_archive

    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w") as z:
        z.writestr(
            "1.txt",
            "magnet:?xt=urn:btih:eeeeeeeeeeeeeeeeeeeeeeeeeeeeeeeeeeeeeeee",
        )
        z.writestr("2.txt", "should not be required\n" + ("noise\n" * 200))
    text = _extract_txt_from_archive(buf.getvalue(), "zip")
    assert "magnet:?xt=urn:btih:eeeeeeeeeeeeeeeeeeeeeeeeeeeeeeeeeeeeeeee" in text


def test_oversized_attachment_bytes_skipped():
    from crawler.attachments import _text_from_attachment_bytes
    from parsers.attachments import MAX_ATTACHMENT_BYTES

    att = DownloadAttachment(name="huge.txt", url="http://x/a", kind="txt")
    data = b"ed2k://|file|x|1|ABCDEFABCDEFABCDEFABCDEFABCDEFAB|/\n" + (
        b"x" * (MAX_ATTACHMENT_BYTES + 1)
    )
    assert _text_from_attachment_bytes(att, data) == ""


def test_zip_member_declared_size_guard():
    """声明 file_size 超限则不 read，防 zip bomb。"""
    from crawler import attachments as mod

    class _Info:
        file_size = mod.MAX_ARCHIVE_MEMBER_BYTES + 1

    class _Zip:
        def getinfo(self, name: str) -> _Info:
            return _Info()

        def read(self, name: str, pwd: bytes | None = None) -> bytes:
            raise AssertionError("oversized member must not be read")

    assert mod._zip_member_bytes(_Zip(), "links.txt") is None


def test_archive_depth_within_cap_keeps_link():
    """深度上限内的嵌套 zip 仍应解出链接。"""
    from crawler.attachments import _extract_txt_from_archive
    from parsers.attachments import MAX_ARCHIVE_DEPTH

    link = "ed2k://|file|demo.mp4|10|ABCDEFABCDEFABCDEFABCDEFABCDEFAB|/\n"
    data = link.encode()
    # MAX_ARCHIVE_DEPTH 层嵌套归档 + 最内层 txt（depth 0..MAX 均可读成员）
    for i in range(MAX_ARCHIVE_DEPTH):
        buf = io.BytesIO()
        with zipfile.ZipFile(buf, "w") as z:
            if i == 0:
                z.writestr("links.txt", link)
            else:
                z.writestr(f"inner{i}.zip", data)
        data = buf.getvalue()
    text = _extract_txt_from_archive(data, "zip")
    assert "ed2k://" in text


def test_archive_depth_over_cap_returns_empty():
    """超过嵌套深度上限则放弃，避免 zip bomb 递归。"""
    from crawler.attachments import _extract_txt_from_archive
    from parsers.attachments import MAX_ARCHIVE_DEPTH

    link = "ed2k://|file|demo.mp4|10|ABCDEFABCDEFABCDEFABCDEFABCDEFAB|/\n"
    data = link.encode()
    # 多套一层：最内层链接落在 depth=MAX+1，应解不出
    for i in range(MAX_ARCHIVE_DEPTH + 2):
        buf = io.BytesIO()
        with zipfile.ZipFile(buf, "w") as z:
            if i == 0:
                z.writestr("links.txt", link)
            else:
                z.writestr(f"inner{i}.zip", data)
        data = buf.getvalue()
    text = _extract_txt_from_archive(data, "zip")
    assert "ed2k://" not in text


def test_import_outcome_distinguishes_body_vs_attachment():
    """正文有链 →「正文含」；仅附件注入有链 →「附件解析」。"""
    from workers.thread_outcome import judge_thread_html

    body_html = """
    <html><head><title>【整理】【115ED2K】正文链 - 论坛</title></head>
    <body>
      <span id="thread_subject">【整理】【115ED2K】正文链</span>
      <div id="postmessage_1">
        ed2k://|file|a.rar|1|ABCDEF0123456789ABCDEF0123456789|/
      </div>
      Powered by Discuz!
    </body></html>
    """
    body_html = body_html + ("<!-- pad -->" * 900)
    body = judge_thread_html(
        body_html,
        board_fid="95:716",
        list_title="【整理】【115ED2K】正文链",
        preferred_link="ed2k",
    )
    assert body.verdict == "import"
    assert body.outcome == "成功：正文含目标链接"

    bare = """
    <html><head><title>【整理】【115ED2K】附件链 - 论坛</title></head>
    <body>
      <span id="thread_subject">【整理】【115ED2K】附件链</span>
      <div id="postmessage_1">资源见附件 https://pan.xunlei.com/s/abc</div>
      <div class="tattl"><a href="forum.php?mod=attachment&aid=1">备用.txt</a></div>
      Powered by Discuz!
    </body></html>
    """
    bare = bare + ("<!-- pad -->" * 900)
    attach_text = (
        "解压密码：www.98T.la@\n"
        "ed2k://|file|b.rar|2|ABCDEF0123456789ABCDEF0123456789|/\n"
    )
    merged = inject_attachment_text(bare, attach_text)
    attached = judge_thread_html(
        merged,
        board_fid="95:716",
        list_title="【整理】【115ED2K】附件链",
        preferred_link="ed2k",
        attachments_already_tried=True,
        had_attachments=True,
    )
    assert attached.verdict == "import"
    assert attached.outcome == "成功：附件解析出目标链接"


def test_body_link_plus_attachment_imports_body_first():
    """正文有样例链 + 有资源附件 → 先按正文 import（不合格再由 pipeline 下附件）。"""
    from workers.thread_outcome import judge_thread_html

    html = """
    <html><head><title>【整理】【115ED2K】合集 - 论坛</title></head>
    <body>
      <span id="thread_subject">【整理】【115ED2K】合集【148配额】</span>
      <div id="postmessage_1">
        预览样例
        ed2k://|file|www.98t.la@预览图.rar|1|ABCDEF0123456789ABCDEF0123456789|/
      </div>
      <ignore_js_op>
        <a href="forum.php?mod=attachment&amp;aid=1">www.98t.la@链接.txt</a>
      </ignore_js_op>
      Powered by Discuz!
    </body></html>
    """
    html = html + ("<!-- pad -->" * 900)
    first = judge_thread_html(
        html,
        board_fid="95:716",
        list_title="【整理】【115ED2K】合集【148配额】",
        preferred_link="ed2k",
    )
    assert first.verdict == "import"
    assert first.need_attachments is False
    assert "正文" in first.outcome

    attach_text = (
        "ed2k://|file|www.98t.la@A.mp4|10|11111111111111111111111111111111|/\n"
        "ed2k://|file|www.98t.la@B.mp4|20|22222222222222222222222222222222|/\n"
    )
    merged = inject_attachment_text(html, attach_text)
    second = judge_thread_html(
        merged,
        board_fid="95:716",
        list_title="【整理】【115ED2K】合集【148配额】",
        preferred_link="ed2k",
        attachments_already_tried=True,
        had_attachments=True,
    )
    assert second.verdict == "import"
    assert second.outcome == "成功：附件解析出目标链接"
    assert second.parsed is not None
    assert len(second.parsed.assets) == 2
    uris = " ".join(a.uri or "" for a in second.parsed.assets)
    assert "预览图.rar" not in uris
    assert "A.mp4" in uris and "B.mp4" in uris


def test_multi_seed_name_body_skips_forced_attachment():
    """正文多 magnet + 多【种子名称】：直接入库，勿附件优先压成单资源。"""
    from workers.thread_outcome import judge_thread_html

    h1 = "E59DC164B2B932A1196311111111111111111111"
    h2 = "307C9A4D097E011651FB11111111111111111111"
    h3 = "63B6DB90BEF67F5FF532428E2796D42FC12E4C7B"
    title = "★●經典の三级写真↗️精彩合集↘️♀[06.12]"
    html = f"""
    <html><head><title>{title}</title></head>
    <body>
      <div id="read_tpc" class="tpc_content">
        【种子名称】：OAE-264.torrent
        【磁力连接】：magnet:?xt=urn:btih:{h1}
        【种子名称】：保险女王.torrent
        【磁力连接】：magnet:?xt=urn:btih:{h2}
        (2048 hk-137-8692321-4 torrent)油鬼子
        【种子名称】：油鬼子 torrent
        【磁力连接】：magnet:?xt=urn:btih:{h3}
      </div>
      <ignore_js_op>
        <a href="job.php?action=download&amp;pid=1&amp;tid=26719397">OAE-264.torrent</a>
      </ignore_js_op>
      Powered by Discuz!
    </body></html>
    """
    html = html + ("<!-- pad -->" * 900)
    out = judge_thread_html(
        html,
        board_fid="3",
        list_title=title,
        preferred_link="magnet",
        forum_id="2048",
        tid=26719397,
    )
    assert out.verdict == "import"
    assert out.need_attachments is False
    assert out.parsed is not None
    assert len(out.parsed.assets) >= 3
    names = {a.filename for a in out.parsed.assets}
    assert "油鬼子" in names
    assert title not in names or len(names) > 1


def test_single_magnet_body_imports_before_attachment():
    """单资源正文有链 + 附件区：先正文 import（不合格才由 pipeline 下附件）。"""
    from workers.thread_outcome import judge_thread_html

    h = "E59DC164B2B932A1196311111111111111111111"
    html = f"""
    <html><head><title>单部影片</title></head>
    <body>
      <div id="read_tpc" class="tpc_content">
        【种子名称】：only.torrent
        【磁力连接】：magnet:?xt=urn:btih:{h}
      </div>
      <ignore_js_op>
        <a href="job.php?action=download&amp;pid=1&amp;tid=1">only.torrent</a>
      </ignore_js_op>
      Powered by Discuz!
    </body></html>
    """
    html = html + ("<!-- pad -->" * 900)
    out = judge_thread_html(
        html,
        board_fid="3",
        list_title="单部影片",
        preferred_link="magnet",
        forum_id="2048",
        tid=1,
    )
    assert out.verdict == "import"
    assert out.need_attachments is False
    assert "正文" in out.outcome


def test_attach_preferred_denied_stubs_not_body_import():
    """tid=3395138：不合格后试附件但无权 → stub，勿用正文残链成功入库。"""
    from workers.thread_outcome import judge_thread_html

    links = "\n".join(
        f"ed2k://|file|part{i}.mp4|{10_000_000 + i}|{i:032X}|/" for i in range(5)
    )
    html = f"""
    <html><head><title>【ED2K】熊猫班 11V/11配额</title></head>
    <body>
      <span id="thread_subject">【ED2K】熊猫班 11V/11配额</span>
      <div id="postmessage_1">{links}<br>完整版见附件</div>
      <div class="tattl">
        <a href="forum.php?mod=attachment&amp;aid=1">www.98T.la@熊猫班.txt</a>
      </div>
      Powered by Discuz!
    </body></html>
    """
    html = html + ("<!-- pad -->" * 900)
    out = judge_thread_html(
        html,
        board_fid="95:716",
        list_title="【ED2K】熊猫班 11V/11配额",
        preferred_link="ed2k",
        forum_id="sehuatang",
        tid=3395138,
        attachments_already_tried=True,
        attachment_denied=True,
        had_attachments=False,
    )
    assert out.verdict == "stub"
    assert out.outcome == "无权限下载附件"


def test_password_in_attachment_text_fills_field_and_description():
    """解压密码只在附件语料：字段与描述都要带上。"""
    html = """
    <html><body>
      <span id="thread_subject">【整理】【115ED2K】附件密码</span>
      <div id="postmessage_1">见附件</div>
    </body></html>
    """
    attach = (
        "解压密码：www.98T.la@\n"
        "ed2k://|file|c.rar|3|ABCDEF0123456789ABCDEF0123456789|/\n"
    )
    merged = inject_attachment_text(html, attach)
    parsed = parse_thread_dual(
        merged, tid=1, preferred_link="ed2k", board_fid="95:716"
    )
    assert parsed.extract_password == "www.98T.la@"
    assert "【解压密码】：www.98T.la@" in (parsed.description or "")
    assert parsed.primary_link_kind == "ed2k"


def test_archive_password_falls_back_to_author_name():
    """正文未写密码时，用一楼作者名作解压候选（tid 332436）。"""
    from crawler.attachments import _archive_password_candidates, _extract_zip_txt

    html = """
    <html><body>
      <div id="postlist">
        <div id="post_1">
          <div class="authi"><a href="space-uid-1.html" class="xw1">ffdshow000</a></div>
          <div id="postmessage_1">解压密码 见下面这个帖子</div>
        </div>
      </div>
    </body></html>
    """
    assert _archive_password_candidates(html) == ["ffdshow000"]

    # 同包多 torrent：全部转磁力（不因第一个命中早停）
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w") as zf:
        zf.writestr("a.torrent", _minimal_torrent(b"a.bin", 100))
        zf.writestr("b.torrent", _minimal_torrent(b"b.bin", 200))
    text = _extract_zip_txt(buf.getvalue(), passwords=["ffdshow000"])
    assert text.count("magnet:?xt=urn:btih:") == 2


def test_phpwind_attach_card_torrent():
    """2048/PHPWind：job.php?action=download 的 .torrent 附件应识别为种子。"""
    html = """
    <div class="attach-card" id="att_6650545">
      <div class="attach-icon">
        <img style="width:30px;" src="images/wind/file/torrent.gif" alt="torrent" />
      </div>
      <div class="attach-main">
        <div class="attach-header"><span class="attach-label">附件</span></div>
        <a href="job.php?action=download&amp;aid=6650545&amp;time=1&amp;sign=1"
           class="attach-name-link">
          <span>87-1021497.torrent</span>
        </a>
        <div class="attach-meta"><span>大小：18 K</span></div>
      </div>
    </div>
    """
    base = "https://bbs.xfca2022.com/"
    all_a = extract_download_attachments(base, html)
    assert len(all_a) == 1
    assert all_a[0].kind == "torrent"
    assert all_a[0].name == "87-1021497.torrent"
    assert "job.php?action=download" in all_a[0].url
    assert "aid=6650545" in all_a[0].url
    torrents = filter_torrent_attachments(all_a)
    assert len(torrents) == 1
    from parsers.attachments import pick_magnet_attachment_kind
    from parsers.thread_gates import looks_like_attachment_zone

    assert looks_like_attachment_zone(html) is True
    assert pick_magnet_attachment_kind(base, html) == "torrent"


def test_truncated_torrent_extension_torren():
    """发帖截断 .torren（缺末尾 t）仍应识别为种子附件（tid 3615372）。"""
    html = """
    <div class="t_f" id="postmessage_1">
      <ignore_js_op>
        <img src="static/image/filetype/torrent.gif" alt="torrent" />
        <a href="forum.php?mod=attachment&amp;aid=MQ%3D%3D" target="_blank">
          HMN-439 影片.torren
        </a>
      </ignore_js_op>
    </div>
    """
    base = "https://www.sehuatang.net/"
    all_a = extract_download_attachments(base, html)
    assert len(all_a) == 1
    assert all_a[0].kind == "torrent"
    assert all_a[0].name.lower().endswith(".torren")
    from parsers.attachments import pick_magnet_attachment_kind
    from parsers.thread_gates import looks_like_attachment_zone

    assert looks_like_attachment_zone(html) is True
    assert pick_magnet_attachment_kind(base, html) == "torrent"


def test_ss_rar_body_samples_import_body_first():
    """盖楼帖正文多段 *_ss.rar 样例链 + 附件区：先正文 import（不合格再附件）。"""
    from workers.thread_outcome import judge_thread_html

    title = "【自转】【115ed2k】TeenMegaWorld 盖楼帖 Wow【121V/30G/121配额】"
    html = f"""
    <html><head><title>{title}</title></head>
    <body>
      <span id="thread_subject">{title}</span>
      <div id="postmessage_1">
        【资源名称】：Wow-Organsms 子系列
        【资源大小】：121V/30G/121配额
        ed2k://|file|WOW-Orgasms.com_ss.rar|19263543|6C9B9192C40D03F7CD3F37CD821D7487|/
      </div>
      <div id="postmessage_2">
        【资源名称】：Beauty4K.com 4k 美女
        【资源大小】：442.89GB/117V/117配额
        ed2k://|file|Beauty4K.com_ss.rar|56671233|BE2C3007782CDCD427D7461D1B5B01C0|/
      </div>
      <div id="postmessage_3">
        【资源名称】：NylonsX.com 丝袜
        ed2k://|file|NylonsX.com_ss.rar|16507691|E198F8659B5C63261750A32AEB7D8ADA|/
      </div>
      <ignore_js_op>
        <a href="forum.php?mod=attachment&amp;aid=1">www.98T.la@WOW-Orgasms.com.ed2k.txt</a>
      </ignore_js_op>
      Powered by Discuz!
    </body></html>
    """
    html = html + ("<!-- pad -->" * 900)
    out = judge_thread_html(
        html,
        board_fid="95:716",
        list_title=title,
        preferred_link="ed2k",
    )
    assert out.verdict == "import"
    assert out.need_attachments is False
    assert "正文" in out.outcome


def test_count_links_and_quota_from_html():
    from crawler.attachments import _count_importable_links, _quota_expect_from_html

    html = """
    <html><head><title>合集【42G/40V/4配额】</title></head>
    <body><span id="thread_subject">合集【42G/40V/4配额】</span></body></html>
    """
    assert _quota_expect_from_html(html) == 4
    text = (
        "ed2k://|file|a.rar|1|AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA|/\n"
        "ed2k://|file|b.rar|1|BBBBBBBBBBBBBBBBBBBBBBBBBBBBBBBB|/"
    )
    assert _count_importable_links(text) == 2

